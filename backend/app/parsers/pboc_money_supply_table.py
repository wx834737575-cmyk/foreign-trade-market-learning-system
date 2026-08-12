from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from io import BytesIO
from urllib.parse import urljoin, urlparse

from bs4 import BeautifulSoup
from openpyxl import load_workbook


PARSER_VERSION = "pboc-money-supply-table-v2"


@dataclass(frozen=True)
class MoneySupplyLinks:
    html_url: str
    workbook_url: str
    pdf_url: str


@dataclass(frozen=True)
class MoneySupplyPoint:
    dataset_code: str
    period: date
    value: Decimal
    methodology_version: str
    note: str | None = None


def _official_url(base_url: str, href: str) -> str:
    url = urljoin(base_url, href)
    if urlparse(url).hostname != "www.pbc.gov.cn":
        raise ValueError("人民银行统计入口包含非官方域名链接")
    return url


def parse_year_links(html: bytes | str, base_url: str) -> dict[int, str]:
    soup = BeautifulSoup(html, "lxml")
    result: dict[int, str] = {}
    for anchor in soup.find_all("a", href=True):
        text = " ".join(anchor.get_text(" ", strip=True).split())
        match = re.fullmatch(r"(20\d{2})年统计数据", text)
        if match:
            result.setdefault(int(match.group(1)), _official_url(base_url, anchor["href"]))
    if not result:
        raise ValueError("未在人民银行统计数据页找到年度入口")
    return result


def parse_money_overview_url(html: bytes | str, base_url: str) -> str:
    soup = BeautifulSoup(html, "lxml")
    for anchor in soup.find_all("a", href=True):
        text = " ".join(anchor.get_text(" ", strip=True).split())
        if text in {"货币统计概览", "货币统计概览 Money and Banking Statistics"}:
            return _official_url(base_url, anchor["href"])
    raise ValueError("年度统计页中未找到货币统计概览入口")


def parse_money_supply_links(html: bytes | str, base_url: str) -> MoneySupplyLinks:
    soup = BeautifulSoup(html, "lxml")
    candidates = []
    for row in soup.find_all("tr"):
        text = " ".join(row.get_text(" ", strip=True).split())
        if "货币供应量" in text and "Money Supply" in text:
            candidates.append(row)
    if not candidates:
        raise ValueError("货币统计概览中未找到货币供应量下载行")
    row = min(candidates, key=lambda item: len(item.get_text(" ", strip=True)))
    links: dict[str, str] = {}
    for anchor in row.find_all("a", href=True):
        url = _official_url(base_url, anchor["href"])
        path = urlparse(url).path.lower()
        if path.endswith((".xlsx", ".xls")):
            links["workbook"] = url
        elif path.endswith((".htm", ".html")):
            links["html"] = url
        elif path.endswith(".pdf"):
            links["pdf"] = url
    missing = sorted({"html", "workbook", "pdf"} - set(links))
    if missing:
        raise ValueError(f"货币供应量下载链接不完整: {', '.join(missing)}")
    return MoneySupplyLinks(
        html_url=links["html"],
        workbook_url=links["workbook"],
        pdf_url=links["pdf"],
    )


def _decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).replace(",", "").strip()
    if not re.fullmatch(r"-?\d+(?:\.\d+)?", text):
        return None
    return Decimal(text)


def _find_row(rows: list[list[object]], label: str) -> int:
    for index, row in enumerate(rows):
        normalized = "".join(str(value or "").replace(" ", "") for value in row[:3])
        if label in normalized:
            return index
    raise ValueError(f"人民银行工作簿缺少数据行: {label}")


def _month_columns(rows: list[list[object]], year: int) -> tuple[int, list[int]]:
    for row_index, row in enumerate(rows):
        text = " ".join(str(value or "") for value in row[:3])
        if "项目" not in text or "Item" not in text:
            continue
        columns: list[int] = []
        for column in range(3, min(len(row), 15)):
            value = _decimal(row[column])
            if value is None:
                continue
            if int(value) != year:
                raise ValueError(f"工作簿月份表头与年度入口不一致: {value} != {year}")
            columns.append(column)
        if len(columns) != 12:
            raise ValueError(f"工作簿月份列数量异常: {len(columns)}")
        return row_index, columns
    raise ValueError("工作簿中未找到项目/月份表头")


def parse_money_supply_workbook(content: bytes, year: int) -> tuple[MoneySupplyPoint, ...]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        worksheet = next(
            (
                sheet
                for sheet in workbook.worksheets
                if "货币供应量" in str(sheet.cell(1, 1).value or "")
            ),
            None,
        )
        if worksheet is None:
            raise ValueError("工作簿标题不是货币供应量")
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    header_index, month_columns = _month_columns(rows, year)
    row_specs = (
        ("CN_M2_BALANCE", "货币和准货币（M2）", "PBOC official money supply table"),
        (
            "CN_M1_BALANCE",
            "货币（M1）",
            "PBOC M1 revised 2025" if year >= 2025 else "PBOC M1 pre-2025 definition",
        ),
        ("CN_M0_BALANCE", "流通中货币（M0）", "PBOC official money supply table"),
    )
    points: list[MoneySupplyPoint] = []
    for dataset_code, label, methodology in row_specs:
        row_index = _find_row(rows[header_index + 1 :], label) + header_index + 1
        for month, column in enumerate(month_columns, start=1):
            value = _decimal(rows[row_index][column])
            if value is not None:
                points.append(
                    MoneySupplyPoint(
                        dataset_code=dataset_code,
                        period=date(year, month, 1),
                        value=value,
                        methodology_version=methodology,
                    )
                )

    if year == 2025:
        note_index = next(
            (
                index
                for index, row in enumerate(rows)
                if "2024年各月末M1可比余额" in "".join(str(value or "") for value in row)
            ),
            None,
        )
        if note_index is None:
            raise ValueError("2025年工作簿缺少M1新口径的2024年回溯说明")
        backcast_header = next(
            (
                index
                for index in range(note_index + 1, len(rows))
                if len(rows[index]) >= 15
                and sum(1 for value in rows[index][3:15] if _decimal(value) is not None) == 12
                and all(int(_decimal(value) or 0) == 2024 for value in rows[index][3:15])
            ),
            None,
        )
        if backcast_header is None:
            raise ValueError("2025年工作簿缺少2024年M1回溯月份表头")
        balance_index = next(
            (
                index
                for index in range(backcast_header + 1, min(backcast_header + 6, len(rows)))
                if "余额（亿元）" in "".join(str(value or "") for value in rows[index][:3])
            ),
            None,
        )
        if balance_index is None:
            raise ValueError("2025年工作簿缺少2024年M1可比余额")
        for month, column in enumerate(range(3, 15), start=1):
            value = _decimal(rows[balance_index][column])
            if value is None:
                raise ValueError(f"2024年M1可比余额缺少{month}月")
            points.append(
                MoneySupplyPoint(
                    dataset_code="CN_M1_BALANCE",
                    period=date(2024, month, 1),
                    value=value,
                    methodology_version="PBOC M1 revised 2025 comparable backcast",
                    note="人民银行在2025年货币供应量表中发布的2024年新口径可比余额",
                )
            )
    return tuple(points)


def parse_money_supply_html(content: bytes | str, year: int) -> dict[tuple[str, date], Decimal]:
    soup = BeautifulSoup(content, "lxml")
    rows = soup.find_all("tr")
    header = next(
        (
            row
            for row in rows
            if "项目" in row.get_text(" ", strip=True)
            and "Item" in row.get_text(" ", strip=True)
            and f"{year}.01" in row.get_text(" ", strip=True)
        ),
        None,
    )
    if header is None:
        raise ValueError("HTML核验表缺少月份表头")
    labels = {
        "CN_M2_BALANCE": "货币和准货币（M2）",
        "CN_M1_BALANCE": "货币（M1）",
        "CN_M0_BALANCE": "流通中货币（M0）",
    }
    result: dict[tuple[str, date], Decimal] = {}
    for dataset_code, label in labels.items():
        row = next((item for item in rows if label in item.get_text(" ", strip=True)), None)
        if row is None:
            raise ValueError(f"HTML核验表缺少数据行: {label}")
        values = [_decimal(cell.get_text(" ", strip=True)) for cell in row.find_all(["th", "td"])[1:]]
        numeric = [value for value in values if value is not None]
        for month, value in enumerate(numeric, start=1):
            result[(dataset_code, date(year, month, 1))] = value
    if not result:
        raise ValueError("HTML核验表没有可用数据")
    return result

