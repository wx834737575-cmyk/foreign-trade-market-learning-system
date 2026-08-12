from io import BytesIO

import pytest
from openpyxl import Workbook

from app.parsers.pboc_money_supply_table import (
    parse_money_overview_url,
    parse_money_supply_html,
    parse_money_supply_links,
    parse_money_supply_workbook,
    parse_year_links,
)


def _workbook_bytes(year: int, *, include_backcast: bool = False) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "货币供应量"
    sheet["A6"] = "项目 Item"
    for month, column in enumerate(range(4, 16), start=1):
        sheet.cell(6, column, year + month / 100)
    sheet["A8"] = "货币和准货币（M2）"
    sheet["B10"] = "货币（M1）"
    sheet["C12"] = "流通中货币（M0）"
    for month, column in enumerate(range(4, 16), start=1):
        sheet.cell(8, column, 3000000 + month)
        sheet.cell(10, column, 1000000 + month)
        sheet.cell(12, column, 100000 + month)
    if include_backcast:
        sheet["A15"] = "注：按可比口径回溯后，2024年各月末M1可比余额和增速分别为："
        for month, column in enumerate(range(4, 16), start=1):
            sheet.cell(17, column, 2024 + month / 100)
            sheet.cell(18, column, 900000 + month)
        sheet["C18"] = "余额（亿元）"
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_discover_official_year_and_download_links() -> None:
    index = '<a href="/2026/index.html">2026年统计数据</a>'
    assert parse_year_links(index, "https://www.pbc.gov.cn/stats/index.html") == {
        2026: "https://www.pbc.gov.cn/2026/index.html"
    }
    year_page = '<a href="money/index.html">货币统计概览</a>'
    overview = parse_money_overview_url(year_page, "https://www.pbc.gov.cn/2026/index.html")
    assert overview == "https://www.pbc.gov.cn/2026/money/index.html"
    downloads = """
    <table><tr><td>货币供应量<br>Money Supply</td>
    <td><a href="money.htm">htm</a></td>
    <td><a href="money.xlsx">xls</a></td>
    <td><a href="money.pdf">pdf</a></td></tr></table>
    """
    links = parse_money_supply_links(downloads, overview)
    assert links.workbook_url.endswith("/2026/money/money.xlsx")


def test_parse_workbook_and_2024_m1_comparable_backcast() -> None:
    points = parse_money_supply_workbook(_workbook_bytes(2025, include_backcast=True), 2025)
    assert len(points) == 48
    june_m2 = next(point for point in points if point.dataset_code == "CN_M2_BALANCE" and point.period.month == 6)
    assert june_m2.value == 3000006
    backcast = [point for point in points if point.period.year == 2024]
    assert len(backcast) == 12
    assert all("backcast" in point.methodology_version for point in backcast)


def test_html_values_must_match_monthly_table() -> None:
    html = """
    <table>
      <tr><td>项目 Item</td><td>2026.01</td><td>2026.02</td></tr>
      <tr><td>货币和准货币（M2）</td><td>10</td><td>11</td></tr>
      <tr><td>货币（M1）</td><td>8</td><td>9</td></tr>
      <tr><td>流通中货币（M0）</td><td>2</td><td>3</td></tr>
    </table>
    """
    values = parse_money_supply_html(html, 2026)
    assert len(values) == 6
    assert str(values[("CN_M1_BALANCE", next(period for code, period in values if code == "CN_M1_BALANCE"))]) in {"8", "9"}


def test_reject_non_official_download_domain() -> None:
    html = """
    <tr><td>货币供应量 Money Supply</td>
    <td><a href="https://example.com/money.htm">htm</a></td>
    <td><a href="money.xlsx">xls</a></td><td><a href="money.pdf">pdf</a></td></tr>
    """
    with pytest.raises(ValueError, match="非官方域名"):
        parse_money_supply_links(html, "https://www.pbc.gov.cn/data/index.html")

